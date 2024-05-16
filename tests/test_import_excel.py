
import copy
import unittest

import openpyxl

import grenml.exceptions
from grenml.excel_converter import XLSParser

TEST_DATA = {
    'Institutions': [
        {
            'Name': 'Test Institution',
            'Latitude': '12.345',
            'Longitude': '54.321',
            'Altitude': '0.0',
            'Address': '123 Main St. Test City TC 1A2 B3C',
            'URL': 'testinstitution.com',
            'Type': 'Connected Institution'
        }
    ],
    'Nodes': [
        {
            'Name': 'Test Node 1',
            'Label': 'TEST-NODE-1',
            'Internal ID': 'TEST-NODE-1',
            'Latitude': '23.456',
            'Longitude': '65.432',
            'Altitude': '0.0',
            'Address': '456 Main St. Test City TC 2B3 C4D',
            'Type': 'Ring Node',
            'Ownership': 'Test Institution',
            'Co-Ownership': 'None'
        },
        {
            'Name': 'Test Node 2',
            'Label': 'TEST-NODE-2',
            'Internal ID': 'TEST-NODE-2',
            'Latitude': '34.567',
            'Longitude': '76.543',
            'Altitude': '0.0',
            'Address': '789 Main St. Test City TC 3C4 D5E',
            'Type': 'Ring Node',
            'Ownership': 'Test Institution',
            'Co-Ownership': 'None'
        }
    ],
    'Links': [
        {
            'Name': 'Test Link',
            'Label': 'None',
            'Internal ID': 'None',

            # The keys in a dictionary must be unique
            # but these two columns have the same name in the spreadsheet.
            'Endpoint (1)': 'Test Node 1',
            'Endpoint (2)': 'Test Node 2',

            'Type': 'terrestrial fibre',
            'Throughput': '1.0',
            'Ownership': 'Test Institution',
            'Co-Ownership': 'None'
        }
    ]
}


TEST_INSTITUTION_FIELD_TO_MODEL_ATTRIBUTE = {
    'Name': 'name',
    'Latitude': 'latitude',
    'Longitude': 'longitude',
    'Altitude': 'altitude',
    'Address': 'address',
    'URL': 'url',
    'Type': 'tag'
}


TEST_NODE_FIELD_TO_MODEL_ATTRIBUTE = {
    'Name': 'name',
    'Label': 'short_name',
    'Internal ID': 'id',
    'Latitude': 'latitude',
    'Longitude': 'longitude',
    'Altitude': 'altitude',
    'Address': 'address',
    'Type': 'tag'
}


TEST_LINK_FIELD_TO_MODEL_ATTRIBUTE = {
    'Name': 'name',
    'Label': 'short_name',
    'Internal ID': 'id',
    'Type': 'tag',
    'Throughput': 'throughput'
}


def make_sheet(workbook, title, items, first_row):
    """
    Creates a sheet in the excel workbook received as argument.
    Items is a list of tuples. Each item provides values to a row in the new sheet.
    First_row defines the vertical position where the table begins.
    """
    sheet = workbook.create_sheet(title=title)
    row = first_row
    for item in items:
        column = 1
        for attribute in item:
            sheet.cell(row=row, column=column, value=item[attribute])
            column += 1
        row += 1


def make_workbook(topology_data):
    """ Creates a workbook using the given topology data. """
    workbook = openpyxl.Workbook()
    first_row = 6
    make_sheet(workbook, 'Directions', '', first_row)
    make_sheet(workbook, 'Institutions', topology_data['Institutions'], first_row)
    make_sheet(workbook, 'Nodes', topology_data['Nodes'], first_row)
    make_sheet(workbook, 'Links', topology_data['Links'], first_row)
    workbook.save('test.xlsx')
    return workbook


def get_by_path(obj, path):
    """
    Obj should be a hierarchical aggregate of objects that provide
    the getitem method (dictionary or list that contain dictionaries
    or lists that contain dictionaries or lists ...).

    The function navigates into obj, taking items from path one by one.

    When path is empty, the object that remains is the value returned.
    """
    result = None
    if len(path) == 0:
        result = obj
    else:
        result = get_by_path(obj[path[0]], path[1:])
    return result


def update_by_path(obj, path, update_fn):
    """
    Drills down into obj by repeatedly using elements in the path.

    Replaces the value remaining when the entire path has been used
    with one obtained by applying update_fn to the original value.

    Returns an updated copy of obj.
    """
    original_value = get_by_path(obj, path)
    obj_copy = copy.deepcopy(obj)
    obj_to_change = get_by_path(obj_copy, path[:-1])
    obj_to_change[path[-1]] = update_fn(original_value)
    return obj_copy


class XLSParserTest(unittest.TestCase):
    def check_excel_to_xml_error(self, test_data, error_class):
        """
        Makes a workbook with the given test data,
        attempts to transform it to xml by using XLSParser.

        Check that the error obtained by calling the parser
        is an instance of the given error class.
        """
        workbook = make_workbook(test_data)
        parser = XLSParser('test_topology', 'test_suite')
        grenml_manager, errors = parser.parse_workbook(workbook)

        self.assertTrue(len(errors) >= 1)
        error = errors[0]
        self.assertTrue(isinstance(error, error_class))

    def test_institution_not_found(self):
        # Using update_by_path here and in tests following
        # to make a change to TEST_DATA that breaks
        # the relationships between the objects in it.
        modified_test_data = update_by_path(
            TEST_DATA, ['Institutions', 0, 'Name'], lambda _: 'New Institution Name'
        )
        self.check_excel_to_xml_error(
            modified_test_data,
            grenml.exceptions.InstitutionNotFoundError,
        )

    def test_node_not_found(self):
        modified_test_data = update_by_path(
            TEST_DATA, ['Nodes', 0, 'Name'], lambda _: 'New Node Name'
        )
        self.check_excel_to_xml_error(
            modified_test_data,
            grenml.exceptions.NodeNotFoundError,
        )

    def test_nodes_with_same_id(self):
        node_1_modified = update_by_path(
            get_by_path(TEST_DATA, ['Nodes', 0]),
            ['Name'],
            lambda _: 'Node 1 Modified'
        )
        modified_test_data = update_by_path(
            TEST_DATA, ['Nodes'], lambda nodes: nodes + [node_1_modified]
        )
        self.check_excel_to_xml_error(
            modified_test_data,
            grenml.exceptions.AttributeIdError,
        )

    def test_nodes_with_same_name(self):
        node_1_modified = update_by_path(
            get_by_path(TEST_DATA, ['Nodes', 0]),
            ['Internal ID'],
            lambda _: 'TEST-NODE-1-duplicate'
        )
        modified_test_data = update_by_path(
            TEST_DATA, ['Nodes'], lambda nodes: nodes + [node_1_modified]
        )
        self.check_excel_to_xml_error(
            modified_test_data,
            grenml.exceptions.MultipleReturnedError,
        )

    def test_bad_latitude(self):
        modified_test_data = update_by_path(
            TEST_DATA, ['Nodes', 0, 'Latitude'], lambda _: '=12.345'
        )
        self.check_excel_to_xml_error(
            modified_test_data,
            ValueError,
        )

    def check_model(self, mapping, test_data, model):
        """
        Verifies that values in test_data (dictionary)
        correspond to values in model (instance of one of the classes
        derived from GRENMLObject).

        The mapping argument provides the correspondence
        between test_data and model.
        """
        for test_data_attribute in mapping:
            test_data_value = test_data[test_data_attribute]

            model_attribute = mapping[test_data_attribute]
            properties = model.additional_properties
            if model_attribute in properties:
                # property values are lists
                model_value = properties[model_attribute][0]
            else:
                model_value = model.__getattribute__(model_attribute)

            self.assertEqual(test_data_value, str(model_value))

    def test_normal_usage(self):
        workbook = make_workbook(TEST_DATA)
        parser = XLSParser('test_topology', 'test_suite')
        grenml_manager, errors = parser.parse_workbook(workbook)
        self.assertTrue(len(errors) == 0)

        for test_institution in TEST_DATA['Institutions']:
            institution_model = grenml_manager.get_institution(
                name=test_institution['Name']
            )
            self.check_model(
                TEST_INSTITUTION_FIELD_TO_MODEL_ATTRIBUTE,
                test_institution,
                institution_model,
            )

        for test_node in TEST_DATA['Nodes']:
            node_model = grenml_manager.get_node(
                name=test_node['Name']
            )
            self.check_model(
                TEST_NODE_FIELD_TO_MODEL_ATTRIBUTE,
                test_node,
                node_model,
            )

        for test_link in TEST_DATA['Links']:
            link_model = grenml_manager.get_link(
                name=test_link['Name']
            )
            self.check_model(
                TEST_LINK_FIELD_TO_MODEL_ATTRIBUTE,
                test_link,
                link_model,
            )
